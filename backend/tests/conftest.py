import os

os.environ["TESTING"] = "1"
os.environ["DATABASE_URL"] = "sqlite:///test.db"

from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

SAMPLES_DIR = Path(__file__).resolve().parent.parent.parent / "samples"
SAMPLE_CSV = SAMPLES_DIR / "Mieterliste_1-Garbe (2).csv"
SAMPLE_BVI = SAMPLES_DIR / "BVI Target Tables.xlsx"


@pytest.fixture
def sample_csv_bytes():
    return SAMPLE_CSV.read_bytes()


@pytest.fixture
def sample_bvi_bytes():
    return SAMPLE_BVI.read_bytes()


def make_test_bvi_xlsx(rows: list[dict]) -> bytes:
    """Generate a minimal BVI G2 XLSX with the given property rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "G2_Property_data"
    # Header rows 1-11 (minimal)
    ws.cell(1, 1, "")
    ws.cell(2, 2, "Range 2: Property data")
    ws.cell(11, 2, "Fund ID")
    ws.cell(11, 5, "Property ID")

    col_map = {
        "bvi_fund_id": 2,
        "property_id": 5,
        "predecessor_id": 6,
        "prop_state": 8,
        "ownership_type": 9,
        "land_ownership": 10,
        "country": 11,
        "region": 12,
        "zip_code": 13,
        "city": 14,
        "street": 15,
        "location_quality": 16,
        "green_building_vendor": 17,
        "green_building_cert": 18,
        "green_building_from": 19,
        "green_building_to": 20,
        "ownership_share": 21,
        "purchase_date": 22,
        "construction_year": 23,
        "risk_style": 25,
        "fair_value": 26,
        "market_net_yield": 28,
        "last_valuation_date": 29,
        "next_valuation_date": 30,
        "plot_size_sqm": 32,
        "debt_property": 49,
        "shareholder_loan": 50,
        "co2_emissions": 114,
        "epc_rating": 135,
        "tech_clear_height": 136,
    }

    for i, row in enumerate(rows):
        row_idx = 12 + i
        ws.cell(row_idx, 3, datetime(2025, 3, 31))  # period
        ws.cell(row_idx, 4, "EUR")  # currency
        for field, col in col_map.items():
            if field in row:
                ws.cell(row_idx, col, row[field])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
