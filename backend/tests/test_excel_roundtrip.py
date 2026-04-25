import os

os.environ.setdefault("TESTING", "1")
os.environ.setdefault("DATABASE_URL", "sqlite:///test.db")

from io import BytesIO

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.config import settings
from app.database import Base, get_db
from app.models.database import MasterDataAudit, PropertyMaster

test_engine = create_engine(
    settings.effective_database_url,
    connect_args={"check_same_thread": False},
)
TestSession = sessionmaker(bind=test_engine)


def override_get_db():
    db = TestSession()
    try:
        yield db
    finally:
        db.close()


@pytest.fixture(autouse=True)
def setup_db():
    Base.metadata.create_all(test_engine)
    from app.api.upload import set_session_factory
    set_session_factory(TestSession)
    yield
    set_session_factory(None)
    Base.metadata.drop_all(test_engine)


@pytest.fixture
def client():
    from app.main import app
    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app, raise_server_exceptions=False) as c:
        yield c
    app.dependency_overrides.clear()


@pytest.fixture
def db():
    session = TestSession()
    yield session
    session.close()


def _upload_xlsx(client, xlsx_bytes, endpoint, **params):
    files = {"file": ("test.xlsx", BytesIO(xlsx_bytes), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    return client.post(f"/api/{endpoint}", files=files, params=params)


def test_export_valid_xlsx(client, db):
    db.add_all([
        PropertyMaster(property_id="1001", city="Essen", country="DE", fair_value=1000000),
        PropertyMaster(property_id="1002", city="Berlin", country="DE"),
    ])
    db.commit()

    resp = client.get("/api/master-data/properties/export")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]

    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    assert "property_id" in headers
    assert "city" in headers
    assert ws.cell(3, headers.index("property_id") + 1).value == "1001"
    assert ws.cell(4, headers.index("property_id") + 1).value == "1002"


def test_roundtrip_no_changes(client, db):
    db.add(PropertyMaster(property_id="1001", city="Essen"))
    db.commit()

    resp = client.get("/api/master-data/properties/export")
    preview = _upload_xlsx(client, resp.content, "master-data/properties/import/preview")
    assert preview.status_code == 200
    data = preview.json()
    assert len(data["diffs"]) == 0


def test_roundtrip_with_changes(client, db):
    db.add(PropertyMaster(property_id="1001", city="Essen"))
    db.commit()

    resp = client.get("/api/master-data/properties/export")
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    city_col = headers.index("city") + 1
    ws.cell(3, city_col, "Dortmund")

    buf = BytesIO()
    wb.save(buf)

    preview = _upload_xlsx(client, buf.getvalue(), "master-data/properties/import/preview")
    data = preview.json()
    assert len(data["diffs"]) == 1
    assert data["diffs"][0]["field"] == "city"
    assert data["diffs"][0]["new_value"] == "Dortmund"
    assert data["diffs"][0]["change_type"] == "update"


def test_apply_fill_gaps(client, db):
    db.add(PropertyMaster(property_id="1001", city="Essen"))
    db.commit()

    resp = client.get("/api/master-data/properties/export")
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]

    city_col = headers.index("city") + 1
    country_col = headers.index("country") + 1
    ws.cell(3, city_col, "Dortmund")
    ws.cell(3, country_col, "DE")

    buf = BytesIO()
    wb.save(buf)

    result = _upload_xlsx(client, buf.getvalue(), "master-data/properties/import/apply", mode="fill_gaps")
    data = result.json()
    assert data["updated"] == 1

    db.expire_all()
    prop = db.query(PropertyMaster).filter(PropertyMaster.property_id == "1001").first()
    assert prop.city == "Essen"
    assert prop.country == "DE"


def test_apply_overwrite(client, db):
    db.add(PropertyMaster(property_id="1001", city="Essen"))
    db.commit()

    resp = client.get("/api/master-data/properties/export")
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]

    city_col = headers.index("city") + 1
    ws.cell(3, city_col, "Dortmund")

    buf = BytesIO()
    wb.save(buf)

    result = _upload_xlsx(client, buf.getvalue(), "master-data/properties/import/apply", mode="overwrite")
    data = result.json()
    assert data["updated"] == 1

    db.expire_all()
    prop = db.query(PropertyMaster).filter(PropertyMaster.property_id == "1001").first()
    assert prop.city == "Dortmund"


def test_apply_creates_new(client, db):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1, "Core / Location")
    ws.cell(2, 1, "property_id")
    ws.cell(2, 2, "city")
    ws.cell(3, 1, "9999")
    ws.cell(3, 2, "New City")

    buf = BytesIO()
    wb.save(buf)

    result = _upload_xlsx(client, buf.getvalue(), "master-data/properties/import/apply")
    data = result.json()
    assert data["created"] == 1

    db.expire_all()
    prop = db.query(PropertyMaster).filter(PropertyMaster.property_id == "9999").first()
    assert prop is not None
    assert prop.city == "New City"


def test_apply_audit_entries(client, db):
    db.add(PropertyMaster(property_id="1001", city="Essen"))
    db.commit()

    resp = client.get("/api/master-data/properties/export")
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb.active
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    country_col = headers.index("country") + 1
    ws.cell(3, country_col, "DE")

    buf = BytesIO()
    wb.save(buf)

    _upload_xlsx(client, buf.getvalue(), "master-data/properties/import/apply", mode="fill_gaps")

    db.expire_all()
    audits = db.query(MasterDataAudit).filter(
        MasterDataAudit.change_source == "excel_import"
    ).all()
    assert len(audits) > 0
    assert any(a.field_name == "country" for a in audits)
