from datetime import date
from io import BytesIO

import openpyxl
import pytest

from app.models.database import (
    CsvUpload,
    DataInconsistency,
    FundMapping,
    PropertyMaster,
    RawRentRoll,
    ReportingPeriod,
    SnapshotFundMapping,
    SnapshotPropertyMaster,
    SnapshotTenantMaster,
    TenantMaster,
)


def _make_upload(db, stichtag=None, status="complete"):
    upload = CsvUpload(
        filename="test.csv",
        status=status,
        stichtag=stichtag or date(2025, 3, 31),
        row_count=10,
        data_row_count=5,
        summary_row_count=2,
    )
    db.add(upload)
    db.commit()
    db.refresh(upload)
    return upload


def _add_data_row(db, upload_id, fund, property_id, tenant, unit_type, area, rent):
    row = RawRentRoll(
        upload_id=upload_id, row_number=10, row_type="data",
        fund=fund, property_id=str(property_id), tenant_name=tenant,
        unit_type=unit_type, area_sqm=area, annual_net_rent=rent,
        monthly_net_rent=rent / 12 if rent else None,
    )
    db.add(row)
    db.commit()
    return row


# ── Period CRUD ────────────────────────────────────────────────────

def test_create_period(client, db):
    upload = _make_upload(db)
    resp = client.post("/api/periods", json={"upload_id": upload.id})
    assert resp.status_code == 201
    data = resp.json()
    assert data["status"] == "draft"
    assert data["stichtag"] == "2025-03-31"


def test_create_period_duplicate(client, db):
    upload = _make_upload(db)
    client.post("/api/periods", json={"upload_id": upload.id})
    resp = client.post("/api/periods", json={"upload_id": upload.id})
    assert resp.status_code == 409


def test_create_period_upload_not_complete(client, db):
    upload = _make_upload(db, status="processing")
    resp = client.post("/api/periods", json={"upload_id": upload.id})
    assert resp.status_code == 400


def test_list_periods(client, db):
    upload = _make_upload(db)
    client.post("/api/periods", json={"upload_id": upload.id})
    resp = client.get("/api/periods")
    assert resp.status_code == 200
    assert len(resp.json()) == 1


def test_get_period(client, db):
    upload = _make_upload(db)
    create_resp = client.post("/api/periods", json={"upload_id": upload.id})
    pid = create_resp.json()["id"]
    resp = client.get(f"/api/periods/{pid}")
    assert resp.status_code == 200
    assert resp.json()["id"] == pid


def test_delete_draft_period(client, db):
    upload = _make_upload(db)
    create_resp = client.post("/api/periods", json={"upload_id": upload.id})
    pid = create_resp.json()["id"]
    resp = client.delete(f"/api/periods/{pid}")
    assert resp.status_code == 200
    assert client.get(f"/api/periods/{pid}").status_code == 404


# ── Finalization ───────────────────────────────────────────────────

def test_finalize_check_clean(client, db):
    upload = _make_upload(db)
    db.add(PropertyMaster(property_id="7042", country="NL", city="Almere",
                          prop_state="HELD_PROPERTY", ownership_type="DIRECT",
                          fair_value=1000000, ownership_share=1.0, street="Main 1"))
    db.commit()

    create_resp = client.post("/api/periods", json={"upload_id": upload.id})
    pid = create_resp.json()["id"]

    resp = client.get(f"/api/periods/{pid}/finalize-check")
    assert resp.status_code == 200
    data = resp.json()
    assert data["can_finalize"] is True
    assert data["blocking_errors"] == 0


def test_finalize_check_with_errors(client, db):
    upload = _make_upload(db)
    db.add(DataInconsistency(
        upload_id=upload.id, category="unmapped_tenant",
        severity="error", description="test", status="open",
        entity_type="tenant", entity_id="T1",
    ))
    db.commit()

    create_resp = client.post("/api/periods", json={"upload_id": upload.id})
    pid = create_resp.json()["id"]

    resp = client.get(f"/api/periods/{pid}/finalize-check")
    data = resp.json()
    assert data["can_finalize"] is False
    assert data["unmapped_tenants"] == 1


def test_finalize_creates_snapshots(client, db):
    upload = _make_upload(db)
    db.add(PropertyMaster(property_id="7042", country="NL", city="Almere"))
    db.add(TenantMaster(tenant_name_canonical="Acme GmbH", bvi_tenant_id="T001"))
    db.add(FundMapping(csv_fund_name="GLIF", bvi_fund_id="GLIFLUF"))
    db.commit()

    create_resp = client.post("/api/periods", json={"upload_id": upload.id})
    pid = create_resp.json()["id"]

    resp = client.post(f"/api/periods/{pid}/finalize")
    assert resp.status_code == 200
    data = resp.json()
    assert data["status"] == "finalized"
    assert data["snapshot_counts"]["properties"] == 1
    assert data["snapshot_counts"]["tenants"] == 1
    assert data["snapshot_counts"]["funds"] == 1

    db.expire_all()
    snaps = db.query(SnapshotPropertyMaster).filter(
        SnapshotPropertyMaster.reporting_period_id == pid
    ).all()
    assert len(snaps) == 1
    assert snaps[0].property_id == "7042"
    assert snaps[0].country == "NL"


def test_finalize_twice_rejected(client, db):
    upload = _make_upload(db)
    create_resp = client.post("/api/periods", json={"upload_id": upload.id})
    pid = create_resp.json()["id"]
    client.post(f"/api/periods/{pid}/finalize")
    resp = client.post(f"/api/periods/{pid}/finalize")
    assert resp.status_code == 400


def test_delete_finalized_rejected(client, db):
    upload = _make_upload(db)
    create_resp = client.post("/api/periods", json={"upload_id": upload.id})
    pid = create_resp.json()["id"]
    client.post(f"/api/periods/{pid}/finalize")
    resp = client.delete(f"/api/periods/{pid}")
    assert resp.status_code == 400


# ── Export ─────────────────────────────────────────────────────────

def test_export_draft(client, db):
    upload = _make_upload(db)
    db.add(FundMapping(csv_fund_name="GLIF", bvi_fund_id="GLIFLUF"))
    db.commit()
    _add_data_row(db, upload.id, "GLIF", "7042", "Tenant A", "Halle", 5000, 120000)

    create_resp = client.post("/api/periods", json={"upload_id": upload.id})
    pid = create_resp.json()["id"]

    resp = client.get(f"/api/periods/{pid}/export")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "DRAFT" in resp.headers["content-disposition"]

    wb = openpyxl.load_workbook(BytesIO(resp.content))
    assert "Z1_Tenants_Leases" in wb.sheetnames
    assert "G2_Property_data" in wb.sheetnames

    ws_z1 = wb["Z1_Tenants_Leases"]
    assert ws_z1.cell(1, 2).value == "PROVISIONAL - NOT FINALIZED"
    assert ws_z1.cell(12, 7).value == "Tenant A"
    assert ws_z1.cell(12, 11).value == 120000

    ws_g2 = wb["G2_Property_data"]
    assert ws_g2.cell(12, 5).value == "7042"
    assert ws_g2.cell(12, 33).value == 5000


def test_export_finalized(client, db):
    upload = _make_upload(db)
    _add_data_row(db, upload.id, "GLIF", "7042", "T", "Halle", 5000, 120000)

    create_resp = client.post("/api/periods", json={"upload_id": upload.id})
    pid = create_resp.json()["id"]
    client.post(f"/api/periods/{pid}/finalize")

    resp = client.get(f"/api/periods/{pid}/export")
    assert resp.status_code == 200
    assert "DRAFT" not in resp.headers["content-disposition"]

    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws_z1 = wb["Z1_Tenants_Leases"]
    assert ws_z1.cell(1, 2).value is None


def test_export_g2_headers(client, db):
    upload = _make_upload(db)
    _add_data_row(db, upload.id, "GLIF", "7042", "T", "Halle", 5000, 120000)

    create_resp = client.post("/api/periods", json={"upload_id": upload.id})
    pid = create_resp.json()["id"]

    resp = client.get(f"/api/periods/{pid}/export")
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws_g2 = wb["G2_Property_data"]

    assert ws_g2.cell(2, 2).value == "Range 2: Property data"
    assert ws_g2.cell(10, 51).value == "Contract and target rents"
    assert ws_g2.cell(11, 2).value == "Fund ID"
    assert ws_g2.cell(11, 33).value == "Rentable area"
    assert ws_g2.cell(11, 51).value == "Contract rent"
    assert ws_g2.cell(11, 144).value == "Reversion"


def test_export_lease_expiry_columns(client, db):
    upload = _make_upload(db)
    _add_data_row(db, upload.id, "GLIF", "7042", "T1", "Halle", 5000, 100000)
    row = db.query(RawRentRoll).filter(RawRentRoll.tenant_name == "T1").first()
    row.lease_end_actual = date(2025, 12, 31)
    db.commit()

    create_resp = client.post("/api/periods", json={"upload_id": upload.id})
    pid = create_resp.json()["id"]

    resp = client.get(f"/api/periods/{pid}/export")
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws_g2 = wb["G2_Property_data"]

    assert ws_g2.cell(12, 100).value == 100000


def test_export_not_found(client):
    resp = client.get("/api/periods/999/export")
    assert resp.status_code == 404
