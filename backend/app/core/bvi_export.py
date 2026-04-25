from __future__ import annotations

from dataclasses import asdict
from datetime import date
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.core.aggregation import aggregate_g2, aggregate_z1
from app.models.database import CsvUpload, ReportingPeriod

HEADER_FONT = Font(bold=True, size=9)
TITLE_FONT = Font(bold=True, size=11)
LIGHT_FILL = PatternFill("solid", fgColor="DCE6F1")

Z1_LABELS = [
    None,  # col A is always empty
    "Fund ID", "Key date", "Currency", "Tenant ID", "DUNS ID",
    "Tenant name", "Industry", "Default probability min",
    "Default probability max", "Net contract rent as of reporting date",
]

Z1_BVI_CODES = [
    None,
    "COMPANY.OBJECT_ID_SENDER", "PERIOD.IDENTIFIER", "CURRENCY",
    "OBJECT_ID_SENDER", "DUNS_ID", "LABEL", "SECTOR",
    "PD_MIN", "PD_MAX", "CONTRACTUAL_RENT_TENANT",
]

Z1_BVI_NUMS = [None, 102, 101, 100, 202, 436, 103, 204, 205, 206, 207]

Z1_TYPES = [
    None, "Alpha-numerical", "Date", "Text", "Alpha-numerical",
    "Numerical", "Text", "Coding", "Numerical", "Numerical", "Numerical",
]

Z1_GROUP_HEADERS = {
    2: "Data set", 4: "Currency", 5: "Tenants", 11: "Rents",
}


G2_LABELS = [
    None,  # A
    "Fund ID", "Key date", "Currency", "Property ID",
    "Hierarchical predecessor ID", "Property name", "Status",
    "Ownership type", "Land Ownership",
    # Address
    "Country", "Region", "Postcode", "City", "Street, address",
    "Quality of location",
    # Green building
    "Green building provider", "Green building certificate",
    "Green building certificate valid as of",
    "Green building certificate valid until",
    # Ownership
    "Percentage ownership", "Acquisition date",
    "Economic construction date", "Main use type", "Risk segment",
    # Valuation
    "Fair market value", "Gross income at arm's length",
    "Property yield", "Date of latest valuation",
    "Scheduled date for next valuation",
    # Floor area
    "Floor area unit of measurement", "Plot Size", "Rentable area",
    "Area Check", "Number of Tenants", "Let area",
    "Office", "Mezzanine", "Industrial (Storage, Warehouse)",
    "Freifläche", "Gastronomy", "Retail", "Hotel", "Rampe",
    "Residential", "Other lettable Area",
    # Parking
    "Number of parking spots", "Number of parking spots - let",
    # Debt
    "Property debt capital", "Property shareholder loans",
    # Rents
    "Contract rent", "rent / sqm", "Targeted net rent",
    # Rent by type
    "Targeted net rent: Office", "Targeted net rent: Mezzanine",
    "Targeted net rent: Industrial, outdoor",
    "Targeted net rent: Industrial (storage, warehouses)",
    "Targeted net rent: Freifläche", "Targeted net rent: gastronomy",
    "Targeted net rent: Retail", "Targeted net rent: Hotel",
    "Targeted net rent: Rampe", "Targeted net rent: Residential",
    "Targeted net rent: Parking", "Targeted net rent: Other",
    # ERV
    "AM ERV: Total", "AM ERV: Office", "AM ERV: Mezzanine",
    "AM ERV: Industrial (storage, warehouses)", "AM ERV: Freifläche",
    "AM ERV: gastronomy", "AM ERV: Retail", "AM ERV: Hotel",
    "AM ERV: Rampe", "AM ERV: Residential", "AM ERV: Parking",
    "AM ERV: Other",
    # Let rent
    "Targeted net rent: Office - let", "Targeted net rent: Mezzanine - let",
    "Targeted net rent: Industrial (storage, warehouses) - let",
    "Targeted net rent: Freifläche - let",
    "Targeted net rent: Gastronomy - let", "Targeted net rent: Retail - let",
    "Targeted net rent: Hotel - let", "Targeted net rent: Rampe - let",
    "Targeted net rent: Residential - let", "Targeted net rent: Parking - let",
    "Targeted net rent: Other - let",
    # Vacant rent
    "Targeted net rent: Office - vacant",
    "Targeted net rent: Mezzanine - vacant",
    "Targeted net rent: Industrial (storage, warehouses) - vacant",
    "Targeted net rent: Freifläche - vacant",
    "Targeted net rent: Gastronomy - vacant",
    "Targeted net rent: Retail - vacant",
    "Targeted net rent: Hotel - vacant", "Targeted net rent: Rampe - vacant",
    "Targeted net rent: Residential - vacant",
    "Targeted net rent: Parking - vacant",
    "Targeted net rent: Other - vacant",
    # Lease expiry
    "Contract rent of leases expiring in year (t)",
    "Contract rent of leases expiring in year (t+1)",
    "Contract rent of leases expiring in year (t+2)",
    "Contract rent of leases expiring in year (t+3)",
    "Contract rent of leases expiring in year (t+4)",
    "Contract rent of leases expiring in year (t+5)",
    "Contract rent of leases expiring in year (t+6)",
    "Contract rent of leases expiring in year (t+7)",
    "Contract rent of leases expiring in year (t+8)",
    "Contract rent of leases expiring in year (t+9)",
    "Contract rent of leases expiring in year (t+10)",
    "Contract rent of open-ended leases",
    "Weighted remaining lease terms",
    # Tenant count
    "Number of tenants",
    # ESG
    "CO2 emissions", "CO2 measurement year",
    "Energy consumption intensity",
    "Energy consumption intensity normalised",
    "Data quality on energy consumption intensity",
    "Energy reference area",
    "Floor area percentage: Office (CRREM)",
    "Floor area percentage: Retail - high street (CRREM)",
    "Floor area percentage: Retail - shopping centre (CRREM)",
    "Floor area percentage: Retail - warehouse (CRREM)",
    "Floor area percentage: Industrial-warehouse (CRREM)",
    "Floor area percentage: Multi-family (CRREM)",
    "Floor area percentage: Single-family (CRREM)",
    "Floor area percentage: Hotel (CRREM)",
    "Floor area percentage: Lodges, leisure, recreation (CRREM)",
    "Floor area percentage: Health (CRREM)",
    "Floor area percentage: Medical office (CRREM)",
    "Exposure to fossil fuels",
    "Exposure to energy-inefficient real estate assets",
    "Total waste volume", "Percentage of recycled waste",
    "Energy performance certificate rating",
    # Tech specs
    "Max. Clear height", "Floor Load capacity", "Loading Docks",
    "Sprinkler System", "Lighting", "Heating", "MAINTENANCE",
    # col 143 empty
    None,
    # Reversion
    "Reversion",
]

G2_GROUP_HEADERS = {
    2: "Data set", 4: "Currency", 5: "ID and WE status",
    11: "Address", 21: "Percentage", 22: "Acquisition date",
    23: "Allocation data", 26: "Survey", 31: "Floor area",
    47: "Parking spots", 49: "Debt capital and SL",
    51: "Contract and target rents",
    54: "Rent by use type (targeted net rent)",
    78: "Rent by use type (targeted net rent) - let",
    89: "Rent by use type (targeted net rent) - vacant",
    100: "Expiring leases by years (headline rent)",
    113: "Number of tenants", 114: "Sustainability",
}

CRREM_KEYS = [
    "office", "retail_high_street", "retail_shopping_centre",
    "retail_warehouse", "industrial_warehouse", "multi_family",
    "single_family", "hotel", "leisure", "health", "medical_office",
]

G2_FIELD_MAP: list[str | None] = [
    None,  # col A
    "fund_id", "stichtag", "currency", "property_id",
    "predecessor_id", "label", "prop_state", "ownership_type",
    "land_ownership",
    "country", "region", "zip_code", "city", "street", "location_quality",
    "green_building_vendor", "green_building_cert",
    "green_building_from", "green_building_to",
    "ownership_share", "purchase_date", "construction_year",
    "use_type_primary", "risk_style",
    "fair_value", "market_rental_value", "market_net_yield",
    "last_valuation_date", "next_valuation_date",
    "area_measure", "plot_size_sqm", "rentable_area",
    "area_check", "tenant_count", "floorspace_let",
    "area_office", "area_mezzanine", "area_industrial", "area_outdoor",
    "area_gastronomy", "area_retail", "area_hotel", "area_ramp",
    "area_residential", "area_other",
    "parking_total", "parking_let",
    "debt_property", "shareholder_loan",
    "contractual_rent", "rent_per_sqm", "gross_potential_income",
    "rent_office", "rent_mezzanine", "rent_industrial_outdoor",
    "rent_industrial", "rent_outdoor", "rent_gastronomy",
    "rent_retail", "rent_hotel", "rent_ramp", "rent_residential",
    "rent_parking", "rent_other",
    "erv_total", "erv_office", "erv_mezzanine", "erv_industrial",
    "erv_outdoor", "erv_gastronomy", "erv_retail", "erv_hotel",
    "erv_ramp", "erv_residential", "erv_parking", "erv_other",
    "let_rent_office", "let_rent_mezzanine", "let_rent_industrial",
    "let_rent_outdoor", "let_rent_gastronomy", "let_rent_retail",
    "let_rent_hotel", "let_rent_ramp", "let_rent_residential",
    "let_rent_parking", "let_rent_other",
    "vacant_rent_office", "vacant_rent_mezzanine",
    "vacant_rent_industrial", "vacant_rent_outdoor",
    "vacant_rent_gastronomy", "vacant_rent_retail",
    "vacant_rent_hotel", "vacant_rent_ramp", "vacant_rent_residential",
    "vacant_rent_parking", "vacant_rent_other",
    # lease expiry: handled specially
    "lease_exp_0", "lease_exp_1", "lease_exp_2", "lease_exp_3",
    "lease_exp_4", "lease_exp_5", "lease_exp_6", "lease_exp_7",
    "lease_exp_8", "lease_exp_9", "lease_exp_10",
    "lease_open_ended", "lease_term_avg",
    "tenant_count_2",
    "co2_emissions", "co2_measurement_year", "energy_intensity",
    "energy_intensity_normalised", "data_quality_energy",
    "energy_reference_area",
    # CRREM: 11 cols (120-130)
    "crrem_office", "crrem_retail_hs", "crrem_retail_sc",
    "crrem_retail_wh", "crrem_industrial_wh", "crrem_multi_family",
    "crrem_single_family", "crrem_hotel", "crrem_leisure",
    "crrem_health", "crrem_medical_office",
    "exposure_fossil_fuels", "exposure_energy_inefficiency",
    "waste_total", "waste_recycled_pct", "epc_rating",
    "tech_clear_height", "tech_floor_load_capacity",
    "tech_loading_docks", "tech_sprinkler", "tech_lighting",
    "tech_heating", "maintenance",
    None,  # col 143
    "reversion",
]


def _write_z1_headers(ws, stichtag: date | None):
    ws.cell(2, 2, "Add-on module 1: Tenants and Leases").font = TITLE_FONT
    if stichtag:
        ws.cell(2, 6, stichtag)
    ws.cell(3, 2, "Item 15 - tenants")
    for c, code in enumerate(Z1_BVI_CODES):
        if code:
            ws.cell(4, c + 1, code).font = HEADER_FONT
    for c, num in enumerate(Z1_BVI_NUMS):
        if num is not None:
            ws.cell(5, c + 1, num)
    for c, t in enumerate(Z1_TYPES):
        if t:
            ws.cell(7, c + 1, t)
    for c, group in Z1_GROUP_HEADERS.items():
        ws.cell(10, c, group).font = HEADER_FONT
        ws.cell(10, c).fill = LIGHT_FILL
    for c, label in enumerate(Z1_LABELS):
        if label:
            ws.cell(11, c + 1, label).font = HEADER_FONT


def _write_g2_headers(ws, stichtag: date | None):
    ws.cell(2, 2, "Range 2: Property data").font = TITLE_FONT
    if stichtag:
        ws.cell(2, 6, stichtag)
    ws.cell(3, 2, "Property data for deriving allocation data")
    for c, group in G2_GROUP_HEADERS.items():
        ws.cell(10, c, group).font = HEADER_FONT
        ws.cell(10, c).fill = LIGHT_FILL
    for c, label in enumerate(G2_LABELS):
        if label:
            ws.cell(11, c + 1, label).font = HEADER_FONT


def generate_bvi_xlsx(
    db: Session,
    upload_id: int,
    stichtag: date | None = None,
    is_draft: bool = False,
) -> bytes:
    z1_rows = aggregate_z1(db, upload_id, stichtag=stichtag)
    g2_rows = aggregate_g2(db, upload_id, stichtag=stichtag)

    wb = openpyxl.Workbook()

    ws_z1 = wb.active
    ws_z1.title = "Z1_Tenants_Leases"
    _write_z1_headers(ws_z1, stichtag)

    if is_draft:
        ws_z1.cell(1, 2, "PROVISIONAL - NOT FINALIZED").font = Font(
            bold=True, color="FF0000", size=12
        )

    for i, row in enumerate(z1_rows, start=12):
        ws_z1.cell(i, 2, row.bvi_fund_id)
        ws_z1.cell(i, 3, stichtag)
        ws_z1.cell(i, 4, "EUR")
        ws_z1.cell(i, 5, row.bvi_tenant_id)
        ws_z1.cell(i, 6, row.property_id)
        ws_z1.cell(i, 7, row.tenant_name)
        ws_z1.cell(i, 8, row.nace_sector)
        ws_z1.cell(i, 9, row.pd_min)
        ws_z1.cell(i, 10, row.pd_max)
        ws_z1.cell(i, 11, row.contractual_rent)

    ws_g2 = wb.create_sheet("G2_Property_data")
    _write_g2_headers(ws_g2, stichtag)

    if is_draft:
        ws_g2.cell(1, 2, "PROVISIONAL - NOT FINALIZED").font = Font(
            bold=True, color="FF0000", size=12
        )

    for i, g in enumerate(g2_rows, start=12):
        d = asdict(g)
        for col_idx, field_name in enumerate(G2_FIELD_MAP):
            if field_name is None:
                continue
            if field_name.startswith("lease_exp_"):
                bucket = field_name.replace("lease_exp_", "")
                val = d.get("lease_expiry", {}).get(bucket, 0)
            elif field_name == "lease_open_ended":
                val = d.get("lease_expiry", {}).get("open_ended", 0)
            elif field_name.startswith("crrem_"):
                crrem = d.get("crrem_floor_areas") or {}
                crrem_key_idx = col_idx - 120
                if 0 <= crrem_key_idx < len(CRREM_KEYS):
                    val = crrem.get(CRREM_KEYS[crrem_key_idx])
                else:
                    val = None
            else:
                val = d.get(field_name)
            if val is not None:
                ws_g2.cell(i, col_idx + 1, val)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
