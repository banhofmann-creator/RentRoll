from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.audit import log_changes, log_creation, snapshot
from app.database import get_db
from app.models.database import DataInconsistency, PropertyMaster

router = APIRouter(tags=["excel-roundtrip"])

FIELD_GROUPS = {
    "Core / Location": [
        "property_id", "fund_csv_name", "predecessor_id", "prop_state",
        "ownership_type", "land_ownership", "country", "region", "zip_code",
        "city", "street", "location_quality",
    ],
    "Green Building": [
        "green_building_vendor", "green_building_cert",
        "green_building_from", "green_building_to",
    ],
    "Financial / Valuation": [
        "ownership_share", "purchase_date", "construction_year",
        "risk_style", "fair_value", "market_net_yield",
        "last_valuation_date", "next_valuation_date",
        "plot_size_sqm", "debt_property", "shareholder_loan",
    ],
    "ESG / Sustainability": [
        "co2_emissions", "co2_measurement_year", "energy_intensity",
        "energy_intensity_normalised", "data_quality_energy",
        "energy_reference_area", "exposure_fossil_fuels",
        "exposure_energy_inefficiency", "waste_total",
        "waste_recycled_pct", "epc_rating",
    ],
    "Technical Specs": [
        "tech_clear_height", "tech_floor_load_capacity",
        "tech_loading_docks", "tech_sprinkler", "tech_lighting",
        "tech_heating", "maintenance",
    ],
}

ALL_FIELDS: list[str] = []
for fields in FIELD_GROUPS.values():
    ALL_FIELDS.extend(fields)

PROPERTY_FIELDS = ALL_FIELDS


def _normalize(val: object) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return str(val.date())
    if isinstance(val, date):
        return str(val)
    if isinstance(val, Decimal):
        return str(float(val))
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    return str(val)


def _resolve_missing_metadata(db: Session, property_id: str):
    db.query(DataInconsistency).filter(
        DataInconsistency.category == "missing_metadata",
        DataInconsistency.entity_id == property_id,
        DataInconsistency.status == "open",
    ).update(
        {
            "status": "resolved",
            "resolution_note": "Auto-resolved: Excel import",
            "resolved_at": datetime.now(timezone.utc),
        },
        synchronize_session=False,
    )


@router.get("/master-data/properties/export")
def export_properties(db: Session = Depends(get_db)):
    properties = db.query(PropertyMaster).order_by(PropertyMaster.property_id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Properties"

    col = 1
    for group_name, fields in FIELD_GROUPS.items():
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(fields) - 1)
        ws.cell(1, col, group_name)
        for i, field in enumerate(fields):
            ws.cell(2, col + i, field)
        col += len(fields)

    for row_idx, prop in enumerate(properties, start=3):
        for col_idx, field in enumerate(ALL_FIELDS, start=1):
            val = getattr(prop, field, None)
            ws.cell(row_idx, col_idx, val)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=properties_export.xlsx"},
    )


@router.post("/master-data/properties/import/preview")
async def import_preview(
    file: UploadFile,
    db: Session = Depends(get_db),
):
    content = await file.read()
    if not content:
        raise HTTPException(400, "Empty file")

    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "Invalid XLSX file")

    ws = wb.active

    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    if "property_id" not in headers:
        raise HTTPException(400, "Missing property_id column in row 2")

    pid_col = headers.index("property_id")
    field_indices = {}
    for i, h in enumerate(headers):
        if h and h in ALL_FIELDS and h != "property_id":
            field_indices[h] = i

    existing_props = {
        p.property_id: p
        for p in db.query(PropertyMaster).all()
    }

    diffs = []
    for row_idx in range(3, ws.max_row + 1):
        pid = ws.cell(row_idx, pid_col + 1).value
        if not pid:
            continue
        pid = str(pid)
        existing = existing_props.get(pid)

        for field, col_idx in field_indices.items():
            new_val = ws.cell(row_idx, col_idx + 1).value
            if new_val is None:
                continue
            new_norm = _normalize(new_val)

            if existing:
                cur_val = getattr(existing, field, None)
                cur_norm = _normalize(cur_val)
                if cur_norm != new_norm:
                    diffs.append({
                        "property_id": pid,
                        "field": field,
                        "current_value": cur_norm or None,
                        "new_value": new_norm,
                        "change_type": "update",
                    })
            else:
                diffs.append({
                    "property_id": pid,
                    "field": field,
                    "current_value": None,
                    "new_value": new_norm,
                    "change_type": "add",
                })

    return {"diffs": diffs, "total_rows": ws.max_row - 2}


@router.post("/master-data/properties/import/apply")
async def import_apply(
    file: UploadFile,
    mode: str = "fill_gaps",
    db: Session = Depends(get_db),
):
    if mode not in ("fill_gaps", "overwrite"):
        raise HTTPException(400, "Mode must be 'fill_gaps' or 'overwrite'")

    content = await file.read()
    if not content:
        raise HTTPException(400, "Empty file")

    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "Invalid XLSX file")

    ws = wb.active
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    if "property_id" not in headers:
        raise HTTPException(400, "Missing property_id column in row 2")

    pid_col = headers.index("property_id")
    field_indices = {}
    for i, h in enumerate(headers):
        if h and h in ALL_FIELDS and h != "property_id":
            field_indices[h] = i

    created = 0
    updated = 0
    skipped = 0

    for row_idx in range(3, ws.max_row + 1):
        pid = ws.cell(row_idx, pid_col + 1).value
        if not pid:
            continue
        pid = str(pid)

        existing = db.query(PropertyMaster).filter(
            PropertyMaster.property_id == pid
        ).first()

        row_data = {}
        for field, col_idx in field_indices.items():
            val = ws.cell(row_idx, col_idx + 1).value
            if val is not None:
                row_data[field] = val

        if existing:
            old = snapshot(existing, PROPERTY_FIELDS)
            changes = {}
            for field, val in row_data.items():
                if mode == "fill_gaps" and getattr(existing, field, None) is not None:
                    continue
                if _normalize(val) != _normalize(getattr(existing, field, None)):
                    setattr(existing, field, val)
                    changes[field] = val
            if changes:
                log_changes(db, "property_master", existing.id, old, changes,
                            change_source="excel_import")
                updated += 1
            else:
                skipped += 1
        else:
            new_prop = PropertyMaster(property_id=pid, **row_data)
            db.add(new_prop)
            db.flush()
            log_creation(db, "property_master", new_prop.id,
                         snapshot(new_prop, PROPERTY_FIELDS),
                         change_source="excel_import")
            _resolve_missing_metadata(db, pid)
            created += 1

    db.commit()

    return {"created": created, "updated": updated, "skipped": skipped}
