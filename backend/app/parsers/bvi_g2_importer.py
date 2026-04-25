from datetime import date, datetime
from io import BytesIO

import openpyxl

SHEET_NAME = "G2_Property_data"
DATA_START_ROW = 12

# G2 column index (1-indexed) → PropertyMaster field name
COL_MAP = {
    5: "property_id",
    6: "predecessor_id",
    8: "prop_state",
    9: "ownership_type",
    10: "land_ownership",
    11: "country",
    12: "region",
    13: "zip_code",
    14: "city",
    15: "street",
    16: "location_quality",
    17: "green_building_vendor",
    18: "green_building_cert",
    19: "green_building_from",
    20: "green_building_to",
    21: "ownership_share",
    22: "purchase_date",
    23: "construction_year",
    25: "risk_style",
    26: "fair_value",
    28: "market_net_yield",
    29: "last_valuation_date",
    30: "next_valuation_date",
    32: "plot_size_sqm",
    49: "debt_property",
    50: "shareholder_loan",
    114: "co2_emissions",
    115: "co2_measurement_year",
    116: "energy_intensity",
    117: "energy_intensity_normalised",
    118: "data_quality_energy",
    119: "energy_reference_area",
    131: "exposure_fossil_fuels",
    132: "exposure_energy_inefficiency",
    133: "waste_total",
    134: "waste_recycled_pct",
    135: "epc_rating",
    136: "tech_clear_height",
    137: "tech_floor_load_capacity",
    138: "tech_loading_docks",
    139: "tech_sprinkler",
    140: "tech_lighting",
    141: "tech_heating",
    142: "maintenance",
}

DATE_FIELDS = {
    "green_building_from", "green_building_to", "purchase_date",
    "last_valuation_date", "next_valuation_date",
}

INT_FIELDS = {"construction_year", "co2_measurement_year", "tech_loading_docks"}

FLOAT_FIELDS = {
    "ownership_share", "fair_value", "market_net_yield", "plot_size_sqm",
    "debt_property", "shareholder_loan", "co2_emissions", "energy_intensity",
    "energy_intensity_normalised", "energy_reference_area",
    "exposure_fossil_fuels", "exposure_energy_inefficiency",
    "waste_total", "waste_recycled_pct", "tech_clear_height",
    "tech_floor_load_capacity",
}

CRREM_COLS = {
    120: "office",
    121: "retail_high_street",
    122: "retail_shopping_centre",
    123: "retail_warehouse",
    124: "industrial_warehouse",
    125: "multi_family",
    126: "single_family",
    127: "hotel",
    128: "leisure",
    129: "health",
    130: "medical_office",
}


def _coerce_value(field: str, raw):
    if raw is None or raw == "":
        return None

    if field in DATE_FIELDS:
        if isinstance(raw, datetime):
            return raw.date()
        if isinstance(raw, date):
            return raw
        return None

    if field in INT_FIELDS:
        try:
            return int(raw)
        except (ValueError, TypeError):
            return None

    if field in FLOAT_FIELDS:
        try:
            return float(raw)
        except (ValueError, TypeError):
            return None

    if field == "property_id":
        return str(int(raw)) if isinstance(raw, (int, float)) else str(raw).strip()

    return str(raw).strip() if raw is not None else None


def parse_bvi_g2(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """Parse BVI G2 sheet and return (properties, warnings).

    Properties are deduplicated by property_id, merging non-null values
    from multiple period rows.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    warnings: list[str] = []

    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    merged: dict[str, dict] = {}
    fund_ids: dict[str, str] = {}

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        pid_raw = ws.cell(row_idx, 5).value
        if pid_raw is None or pid_raw == "":
            continue

        pid = str(int(pid_raw)) if isinstance(pid_raw, (int, float)) else str(pid_raw).strip()
        if not pid:
            continue

        bvi_fund_id = ws.cell(row_idx, 2).value
        if bvi_fund_id and pid not in fund_ids:
            fund_ids[pid] = str(bvi_fund_id).strip()

        row_data: dict = {"property_id": pid}

        for col_idx, field in COL_MAP.items():
            if field == "property_id":
                continue
            raw = ws.cell(row_idx, col_idx).value
            val = _coerce_value(field, raw)
            if val is not None:
                row_data[field] = val

        crrem = {}
        for col_idx, crrem_key in CRREM_COLS.items():
            raw = ws.cell(row_idx, col_idx).value
            if raw is not None:
                try:
                    crrem[crrem_key] = float(raw)
                except (ValueError, TypeError):
                    pass
        if crrem:
            row_data["crrem_floor_areas_json"] = crrem

        if pid in merged:
            for k, v in row_data.items():
                if k == "property_id":
                    continue
                if k == "crrem_floor_areas_json":
                    existing = merged[pid].get("crrem_floor_areas_json", {})
                    if existing:
                        for ck, cv in v.items():
                            if ck not in existing:
                                existing[ck] = cv
                        merged[pid]["crrem_floor_areas_json"] = existing
                    else:
                        merged[pid]["crrem_floor_areas_json"] = v
                elif merged[pid].get(k) is None:
                    merged[pid][k] = v
        else:
            merged[pid] = row_data

    wb.close()

    result = list(merged.values())
    for pid, fid in fund_ids.items():
        if pid in merged:
            merged[pid]["_bvi_fund_id"] = fid

    return result, warnings
